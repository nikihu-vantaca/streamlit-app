#!/usr/bin/env python3
"""
Comprehensive Streamlit Dashboard for Ticket Evaluation Analysis
"""

import streamlit as st
import sqlite3
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import toml
from langsmith import Client
import json
from collections import defaultdict
import re
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(
    page_title="Ticket Evaluation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #2c3e50;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .info-box {
        background-color: #e8f4fd;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #3498db;
    }
    .tab-content {
        padding: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def get_api_key():
    """Get API key from Streamlit secrets"""
    try:
        return st.secrets["langsmith"]["api_key"]
    except:
        return None

def is_valid_date(date_string):
    """Check if a string is a valid date in YYYY-MM-DD format"""
    if not date_string:
        return False
    
    # Check if it matches the pattern YYYY-MM-DD
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_string):
        return False
    
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def fix_malformed_dates():
    """Fix malformed dates in the database"""
    conn = init_database()
    cursor = conn.cursor()
    
    # Get all records with malformed dates
    cursor.execute('SELECT id, date, experiment_name FROM ticket_evaluations')
    records = cursor.fetchall()
    
    fixed_count = 0
    for record_id, date, experiment_name in records:
        if not is_valid_date(date):
            # Try to extract proper date from experiment name
            new_date = extract_date_from_experiment_name(experiment_name)
            if new_date and is_valid_date(new_date):
                cursor.execute('UPDATE ticket_evaluations SET date = ? WHERE id = ?', (new_date, record_id))
                fixed_count += 1
    
    conn.commit()
    conn.close()
    
    if fixed_count > 0:
        st.success(f"✅ Fixed {fixed_count} malformed dates in the database!")

def extract_date_from_experiment_name(experiment_name):
    """Extract date from experiment name"""
    if not experiment_name:
        return None
    
    try:
        # Try to extract date from experiment name
        # Format: implementation-evaluation-2025-08-15-6e065ee8
        parts = experiment_name.split('-')
        if len(parts) >= 6:
            # The format is: implementation-evaluation-2025-08-15-6e065ee8
            # So parts[2] = "2025" (year), parts[3] = "08" (month), parts[4] = "15" (day)
            year, month, day = parts[2], parts[3], parts[4]
            # Validate the parts are actually numbers
            if year.isdigit() and month.isdigit() and day.isdigit():
                return f"{year}-{month}-{day}"
    except:
        pass
    
    return None

def init_database():
    """Initialize database connection and create tables if needed"""
    conn = sqlite3.connect('ticket_data.db')
    cursor = conn.cursor()
    
    # Create table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ticket_evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            ticket_type TEXT NOT NULL,
            quality TEXT,
            comment TEXT,
            evaluation_key TEXT,
            experiment_name TEXT,
            start_time TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    return conn

def fetch_and_store_latest_data():
    """Fetch latest data from LangSmith and store in database"""
    api_key = get_api_key()
    if not api_key:
        st.error("❌ No API key found. Please check your configuration.")
        return False
    
    try:
        client = Client(api_key=api_key)
        
        # Get the latest timestamp from database
        conn = init_database()
        cursor = conn.cursor()
        cursor.execute('SELECT MAX(start_time) FROM ticket_evaluations')
        latest_time = cursor.fetchone()[0]
        
        if latest_time:
            start_date = datetime.fromisoformat(latest_time.replace('Z', '+00:00'))
        else:
            # If no data, start from a reasonable date
            start_date = datetime.now() - timedelta(days=30)
        
        end_date = datetime.now()
        
        st.info(f"🔄 Fetching data from {start_date.strftime('%Y-%m-%d %H:%M')} to {end_date.strftime('%Y-%m-%d %H:%M')}")
        
        # Fetch runs with pagination to avoid rate limits
        all_runs = []
        batch_size = 100
        total_processed = 0
        max_batches = 50  # Limit to avoid overwhelming
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for batch_num in range(max_batches):
            status_text.text(f"Processing batch {batch_num + 1}/{max_batches}...")
            
            try:
                runs = client.list_runs(
                    project_name="evaluators",
                    start_time=start_date,
                    end_time=end_date,
                    limit=batch_size
                )
                
                batch_runs = list(runs)
                if not batch_runs:
                    break
                
                all_runs.extend(batch_runs)
                total_processed += len(batch_runs)
                
                # Process this batch
                process_runs_batch(batch_runs, conn, cursor)
                
                progress_bar.progress((batch_num + 1) / max_batches)
                
                if len(batch_runs) < batch_size:
                    break
                    
            except Exception as e:
                st.error(f"Error in batch {batch_num + 1}: {e}")
                break
        
        conn.commit()
        conn.close()
        
        st.success(f"✅ Successfully processed {total_processed} runs!")
        return True
        
    except Exception as e:
        st.error(f"❌ Error fetching data: {e}")
        return False

def process_runs_batch(runs_batch, conn, cursor):
    """Process a batch of runs and store in database"""
    for run in runs_batch:
        if not run.outputs or run.name != "detailed_similarity_evaluator":
            continue
        
        # Extract experiment name
        experiment = None
        if hasattr(run, "metadata") and run.metadata and isinstance(run.metadata, dict):
            experiment = run.metadata.get("experiment")
        
        if not experiment:
            continue
        
        # Determine ticket type based on experiment name
        ticket_type = determine_ticket_type(experiment, run.start_time)
        
        # Process evaluation output
        output = run.outputs
        if isinstance(output, dict):
            result = output
        elif isinstance(output, str):
            try:
                result = json.loads(output)
            except:
                continue
        else:
            continue
        
        # Extract evaluation details
        quality = result.get("quality")
        comment = result.get("comment")
        
        # Determine quality category
        quality_category = categorize_quality(quality, comment)
        
        # Extract date from experiment name or use run start time
        date = extract_date_from_experiment(experiment, run.start_time)
        
        # Generate unique ticket ID
        ticket_id = hash(f"{experiment}_{run.id}") % 1000000
        
        # Insert into database
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO ticket_evaluations 
                (ticket_id, date, ticket_type, quality, comment, evaluation_key, experiment_name, start_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                ticket_id,
                date,
                ticket_type,
                quality_category,
                comment,
                'detailed_similarity_evaluator',
                experiment,
                run.start_time.isoformat() if run.start_time else None
            ))
        except Exception as e:
            continue

def determine_ticket_type(experiment, start_time):
    """Determine ticket type based on experiment name and date"""
    # Date cutoff for grouped vs ungrouped evaluation
    cutoff_date = datetime(2025, 8, 15)
    
    if start_time and start_time < cutoff_date:
        # Ungrouped evaluation (pre-August 15, 2025)
        if "zendesk-evaluation-" in experiment:
            # Determine type based on comment content or other logic
            return "homeowner"  # Default for zendesk evaluations
    else:
        # Grouped evaluation (post-August 15, 2025)
        if "implementation-evaluation-" in experiment:
            return "implementation"
        elif "homeowner-pay-evaluation-" in experiment:
            return "homeowner"
        elif "management-pay-evaluation-" in experiment:
            return "management"
    
    return "unknown"

def categorize_quality(quality, comment):
    """Categorize quality based on evaluation output"""
    if quality == "copy_paste":
        return "copy_paste"
    elif quality == "low_quality":
        return "low_quality"
    elif quality == "high_quality":
        return "high_quality"
    elif comment == "empty_bot_answer":
        return "skipped"
    elif comment and "management_company_ticket" in comment:
        return "skipped"
    elif comment and "empty_human_answer" in comment:
        return "skipped"
    else:
        return "unknown"

def extract_date_from_experiment(experiment, start_time):
    """Extract date from experiment name or use start time"""
    # First try to extract from experiment name
    extracted_date = extract_date_from_experiment_name(experiment)
    if extracted_date and is_valid_date(extracted_date):
        return extracted_date
    
    # Fallback to start time
    if start_time:
        return start_time.strftime('%Y-%m-%d')
    
    return datetime.now().strftime('%Y-%m-%d')

def get_evaluation_data(start_date, end_date):
    """Get evaluation data from database for the specified date range"""
    conn = init_database()
    
    query = '''
        SELECT date, ticket_type, quality, COUNT(*) as count
        FROM ticket_evaluations 
        WHERE date >= ? AND date <= ? AND date LIKE '____-__-__'
        GROUP BY date, ticket_type, quality
        ORDER BY date, ticket_type, quality
    '''
    
    df = pd.read_sql_query(query, conn, params=[start_date, end_date])
    conn.close()
    
    return df

def get_evaluation_breakdown_by_type(start_date, end_date):
    """Get detailed breakdown by ticket type"""
    conn = init_database()
    
    query = '''
        SELECT 
            ticket_type,
            quality,
            COUNT(*) as count
        FROM ticket_evaluations 
        WHERE date >= ? AND date <= ? AND date LIKE '____-__-__'
        GROUP BY ticket_type, quality
        ORDER BY ticket_type, quality
    '''
    
    df = pd.read_sql_query(query, conn, params=[start_date, end_date])
    conn.close()
    
    return df

def get_available_dates():
    """Get available dates from database (only valid dates)"""
    conn = init_database()
    cursor = conn.cursor()
    
    # Only get valid dates
    cursor.execute('''
        SELECT MIN(date), MAX(date) 
        FROM ticket_evaluations 
        WHERE date LIKE '____-__-__'
    ''')
    
    min_date, max_date = cursor.fetchone()
    conn.close()
    
    return min_date, max_date

def get_daily_breakdown_data(start_date=None, end_date=None):
    """Get detailed daily breakdown data from database (same as create_daily_breakdown_spreadsheet.py)"""
    conn = sqlite3.connect('ticket_data.db')
    
    if start_date and end_date:
        # Get data with valid dates and date range filter
        query = '''
            SELECT 
                date,
                ticket_type,
                quality,
                comment,
                experiment_name,
                COUNT(*) as count
            FROM ticket_evaluations 
            WHERE date LIKE '____-__-__' AND date >= ? AND date <= ?
            GROUP BY date, ticket_type, quality, comment, experiment_name
            ORDER BY date, ticket_type, quality
        '''
        df = pd.read_sql_query(query, conn, params=[start_date, end_date])
    else:
        # Get all data with valid dates
        query = '''
            SELECT 
                date,
                ticket_type,
                quality,
                comment,
                experiment_name,
                COUNT(*) as count
            FROM ticket_evaluations 
            WHERE date LIKE '____-__-__'
            GROUP BY date, ticket_type, quality, comment, experiment_name
            ORDER BY date, ticket_type, quality
        '''
        df = pd.read_sql_query(query, conn)
    
    conn.close()
    
    return df

def create_detailed_breakdown_spreadsheet(start_date=None, end_date=None):
    """Create a comprehensive daily breakdown spreadsheet (same as create_daily_breakdown_spreadsheet.py)"""
    # Get the raw data
    df = get_daily_breakdown_data(start_date, end_date)
    
    if df.empty:
        return None, None
    
    # Create a comprehensive breakdown
    breakdown_data = []
    
    # Get unique dates
    dates = sorted(df['date'].unique())
    
    for date in dates:
        date_data = df[df['date'] == date]
        
        # Overall summary for this date
        total_tickets = date_data['count'].sum()
        
        # Breakdown by ticket type
        for ticket_type in ['implementation', 'homeowner', 'management']:
            type_data = date_data[date_data['ticket_type'] == ticket_type]
            type_total = type_data['count'].sum()
            
            if type_total > 0:
                # Quality breakdown for this ticket type
                for quality in ['high_quality', 'low_quality', 'copy_paste', 'skipped', 'unknown']:
                    quality_data = type_data[type_data['quality'] == quality]
                    quality_count = quality_data['count'].sum()
                    
                    if quality_count > 0:
                        breakdown_data.append({
                            'Date': date,
                            'Ticket_Type': ticket_type.title(),
                            'Quality': quality.replace('_', ' ').title(),
                            'Count': quality_count,
                            'Percentage_of_Type': round((quality_count / type_total) * 100, 1),
                            'Percentage_of_Total': round((quality_count / total_tickets) * 100, 1),
                            'Total_Type_Tickets': type_total,
                            'Total_Daily_Tickets': total_tickets
                        })
        
        # Add summary row for this date
        breakdown_data.append({
            'Date': date,
            'Ticket_Type': 'TOTAL',
            'Quality': 'ALL',
            'Count': total_tickets,
            'Percentage_of_Type': 100.0,
            'Percentage_of_Total': 100.0,
            'Total_Type_Tickets': total_tickets,
            'Total_Daily_Tickets': total_tickets
        })
    
    # Create DataFrame
    breakdown_df = pd.DataFrame(breakdown_data)
    
    # Add additional summary sheets
    summary_data = []
    
    # Overall summary by ticket type
    overall_by_type = df.groupby('ticket_type')['count'].sum().reset_index()
    for _, row in overall_by_type.iterrows():
        summary_data.append({
            'Summary_Type': 'By_Ticket_Type',
            'Category': row['ticket_type'].title(),
            'Total_Count': row['count'],
            'Percentage': round((row['count'] / overall_by_type['count'].sum()) * 100, 1)
        })
    
    # Overall summary by quality
    overall_by_quality = df.groupby('quality')['count'].sum().reset_index()
    for _, row in overall_by_quality.iterrows():
        summary_data.append({
            'Summary_Type': 'By_Quality',
            'Category': row['quality'].replace('_', ' ').title(),
            'Total_Count': row['count'],
            'Percentage': round((row['count'] / overall_by_quality['count'].sum()) * 100, 1)
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    return breakdown_df, summary_df

def create_quality_distribution_chart(df):
    """Create quality distribution chart"""
    if df.empty:
        return go.Figure()
    
    # Pivot data for stacked bar chart
    pivot_df = df.pivot_table(
        index='date', 
        columns='quality', 
        values='count', 
        fill_value=0
    )
    
    fig = go.Figure()
    
    colors = {
        'high_quality': '#2ecc71',
        'low_quality': '#e74c3c',
        'copy_paste': '#f39c12',
        'skipped': '#95a5a6',
        'unknown': '#34495e'
    }
    
    for quality in pivot_df.columns:
        fig.add_trace(go.Bar(
            name=quality.replace('_', ' ').title(),
            x=pivot_df.index,
            y=pivot_df[quality],
            marker_color=colors.get(quality, '#3498db')
        ))
    
    fig.update_layout(
        title="Quality Distribution Over Time",
        xaxis_title="Date",
        yaxis_title="Number of Tickets",
        barmode='stack',
        height=400
    )
    
    return fig

def create_ticket_type_chart(df):
    """Create ticket type distribution chart"""
    if df.empty:
        return go.Figure()
    
    # Aggregate by ticket type
    type_counts = df.groupby('ticket_type')['count'].sum().reset_index()
    
    fig = px.pie(
        type_counts, 
        values='count', 
        names='ticket_type',
        title="Distribution by Ticket Type",
        color_discrete_map={
            'implementation': '#3498db',
            'homeowner': '#2ecc71',
            'management': '#e74c3c'
        }
    )
    
    fig.update_layout(height=400)
    return fig

def create_evaluation_quality_chart(df):
    """Create evaluation quality breakdown by ticket type"""
    if df.empty:
        return go.Figure()
    
    # Pivot data for stacked bar chart
    pivot_df = df.pivot_table(
        index='ticket_type', 
        columns='quality', 
        values='count', 
        fill_value=0
    )
    
    fig = go.Figure()
    
    colors = {
        'high_quality': '#2ecc71',
        'low_quality': '#e74c3c',
        'copy_paste': '#f39c12',
        'skipped': '#95a5a6',
        'unknown': '#34495e'
    }
    
    for quality in pivot_df.columns:
        fig.add_trace(go.Bar(
            name=quality.replace('_', ' ').title(),
            x=pivot_df.index,
            y=pivot_df[quality],
            marker_color=colors.get(quality, '#3498db')
        ))
    
    fig.update_layout(
        title="Evaluation Quality by Ticket Type",
        xaxis_title="Ticket Type",
        yaxis_title="Number of Tickets",
        barmode='stack',
        height=400
    )
    
    return fig

def create_daily_breakdown_chart(breakdown_df):
    """Create daily breakdown visualization"""
    if breakdown_df.empty:
        return go.Figure()
    
    # Filter out TOTAL rows for visualization
    viz_df = breakdown_df[breakdown_df['Ticket_Type'] != 'TOTAL']
    
    # Create pivot table for visualization
    pivot_df = viz_df.pivot_table(
        index='Date',
        columns=['Ticket_Type', 'Quality'],
        values='Count',
        fill_value=0
    )
    
    # Flatten column names
    pivot_df.columns = [f"{col[0]}_{col[1]}" for col in pivot_df.columns]
    
    fig = go.Figure()
    
    colors = {
        'Implementation_High Quality': '#2ecc71',
        'Implementation_Low Quality': '#e74c3c',
        'Implementation_Copy Paste': '#f39c12',
        'Implementation_Skipped': '#95a5a6',
        'Homeowner_High Quality': '#27ae60',
        'Homeowner_Low Quality': '#c0392b',
        'Homeowner_Copy Paste': '#d68910',
        'Homeowner_Skipped': '#7f8c8d',
        'Management_High Quality': '#2980b9',
        'Management_Low Quality': '#8e44ad',
        'Management_Copy Paste': '#e67e22',
        'Management_Skipped': '#34495e'
    }
    
    for col in pivot_df.columns:
        fig.add_trace(go.Bar(
            name=col.replace('_', ' - '),
            x=pivot_df.index,
            y=pivot_df[col],
            marker_color=colors.get(col, '#3498db')
        ))
    
    fig.update_layout(
        title="Daily Breakdown by Ticket Type and Quality",
        xaxis_title="Date",
        yaxis_title="Number of Tickets",
        barmode='stack',
        height=500,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    return fig

def download_excel_file(breakdown_df, summary_df):
    """Create and download Excel file"""
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Create breakdown sheet
    ws1 = wb.create_sheet("Daily_Breakdown")
    for r in dataframe_to_rows(breakdown_df, index=False, header=True):
        ws1.append(r)
    
    # Create summary sheet
    ws2 = wb.create_sheet("Summary")
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws2.append(r)
    
    # Create pivot tables for better analysis
    # Pivot by date and ticket type
    pivot_type = breakdown_df[breakdown_df['Ticket_Type'] != 'TOTAL'].pivot_table(
        index='Date', 
        columns='Ticket_Type', 
        values='Count', 
        aggfunc='sum',
        fill_value=0
    )
    
    ws3 = wb.create_sheet("Pivot_By_Type")
    for r in dataframe_to_rows(pivot_type, index=True, header=True):
        ws3.append(r)
    
    # Pivot by date and quality
    pivot_quality = breakdown_df[breakdown_df['Ticket_Type'] != 'TOTAL'].pivot_table(
        index='Date', 
        columns='Quality', 
        values='Count', 
        aggfunc='sum',
        fill_value=0
    )
    
    ws4 = wb.create_sheet("Pivot_By_Quality")
    for r in dataframe_to_rows(pivot_quality, index=True, header=True):
        ws4.append(r)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def main():
    """Main Streamlit application"""
    
    # Header
    st.markdown('<h1 class="main-header">📊 Ticket Evaluation Dashboard</h1>', unsafe_allow_html=True)
    
    # Create tabs
    tab1, tab2 = st.tabs(["📈 Overview Dashboard", "📊 Daily Breakdown Analysis"])
    
    # Sidebar (shared across tabs)
    st.sidebar.title("⚙️ Dashboard Controls")
    
    # Fix malformed dates button
    st.sidebar.subheader("🔧 Database Maintenance")
    if st.sidebar.button("Fix Malformed Dates"):
        with st.spinner("Fixing malformed dates..."):
            fix_malformed_dates()
            st.rerun()
    
    # Date range selection
    st.sidebar.subheader("📅 Date Range")
    
    # Get available dates from database (only valid dates)
    min_date, max_date = get_available_dates()
    
    if min_date and max_date:
        try:
            start_date_input = st.sidebar.date_input(
                "Start Date",
                value=datetime.strptime(min_date, '%Y-%m-%d').date(),
                min_value=datetime.strptime(min_date, '%Y-%m-%d').date(),
                max_value=datetime.strptime(max_date, '%Y-%m-%d').date()
            )
            
            end_date_input = st.sidebar.date_input(
                "End Date",
                value=datetime.strptime(max_date, '%Y-%m-%d').date(),
                min_value=datetime.strptime(min_date, '%Y-%m-%d').date(),
                max_value=datetime.strptime(max_date, '%Y-%m-%d').date()
            )
            
            # Handle tuple return from date_input
            start_date = start_date_input[0] if isinstance(start_date_input, tuple) else start_date_input
            end_date = end_date_input[0] if isinstance(end_date_input, tuple) else end_date_input
            
            # Ensure dates are not None
            if start_date is None:
                start_date = datetime.now().date() - timedelta(days=30)
            if end_date is None:
                end_date = datetime.now().date()
            
        except (ValueError, TypeError):
            # Fallback if date parsing fails
            start_date_input = st.sidebar.date_input("Start Date", value=datetime.now().date() - timedelta(days=30))
            end_date_input = st.sidebar.date_input("End Date", value=datetime.now().date())
            start_date = start_date_input[0] if isinstance(start_date_input, tuple) else start_date_input
            end_date = end_date_input[0] if isinstance(end_date_input, tuple) else end_date_input
            
            # Ensure dates are not None
            if start_date is None:
                start_date = datetime.now().date() - timedelta(days=30)
            if end_date is None:
                end_date = datetime.now().date()
    else:
        start_date_input = st.sidebar.date_input("Start Date", value=datetime.now().date() - timedelta(days=30))
        end_date_input = st.sidebar.date_input("End Date", value=datetime.now().date())
        start_date = start_date_input[0] if isinstance(start_date_input, tuple) else start_date_input
        end_date = end_date_input[0] if isinstance(end_date_input, tuple) else end_date_input
        
        # Ensure dates are not None
        if start_date is None:
            start_date = datetime.now().date() - timedelta(days=30)
        if end_date is None:
            end_date = datetime.now().date()
    
    # Sync button
    st.sidebar.subheader("🔄 Data Sync")
    if st.sidebar.button("Sync Latest Data"):
        with st.spinner("Syncing data from LangSmith..."):
            success = fetch_and_store_latest_data()
            if success:
                st.sidebar.success("✅ Sync completed!")
                st.rerun()
    
    # Database info
    st.sidebar.subheader("💾 Database Info")
    conn = init_database()
    cursor = conn.cursor()
    
    # Total records
    cursor.execute('SELECT COUNT(*) FROM ticket_evaluations')
    total_records = cursor.fetchone()[0]
    st.sidebar.metric("Total Records", f"{total_records:,}")
    
    # Valid records (with proper dates)
    cursor.execute('SELECT COUNT(*) FROM ticket_evaluations WHERE date LIKE "____-__-__"')
    valid_records = cursor.fetchone()[0]
    st.sidebar.metric("Valid Records", f"{valid_records:,}")
    
    # Records in selected range
    cursor.execute('SELECT COUNT(*) FROM ticket_evaluations WHERE date >= ? AND date <= ? AND date LIKE "____-__-__"', 
                  [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')])
    range_records = cursor.fetchone()[0]
    st.sidebar.metric("Records in Range", f"{range_records:,}")
    
    # Latest sync
    cursor.execute('SELECT MAX(created_at) FROM ticket_evaluations')
    latest_sync = cursor.fetchone()[0]
    if latest_sync:
        st.sidebar.info(f"Last Sync: {latest_sync}")
    
    conn.close()
    
    # Tab 1: Overview Dashboard
    with tab1:
        st.markdown('<h2 class="section-header">📈 Overview Metrics</h2>', unsafe_allow_html=True)
        
        # Get data for the selected date range
        df = get_evaluation_data(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
        
        if not df.empty:
            # Key metrics
            col1, col2, col3, col4 = st.columns(4)
            
            total_tickets = df['count'].sum()
            with col1:
                st.metric("Total Tickets", f"{total_tickets:,}")
            
            # Quality breakdown
            quality_counts = df.groupby('quality')['count'].sum()
            high_quality_pct = (quality_counts.get('high_quality', 0) / total_tickets * 100) if total_tickets > 0 else 0
            with col2:
                st.metric("High Quality %", f"{high_quality_pct:.1f}%")
            
            copy_paste_pct = (quality_counts.get('copy_paste', 0) / total_tickets * 100) if total_tickets > 0 else 0
            with col3:
                st.metric("Copy-Paste %", f"{copy_paste_pct:.1f}%")
            
            skipped_pct = (quality_counts.get('skipped', 0) / total_tickets * 100) if total_tickets > 0 else 0
            with col4:
                st.metric("Skipped %", f"{skipped_pct:.1f}%")
            
            # Charts
            st.markdown('<h2 class="section-header">📊 Visualizations</h2>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                fig_quality = create_quality_distribution_chart(df)
                st.plotly_chart(fig_quality, use_container_width=True)
            
            with col2:
                fig_type = create_ticket_type_chart(df)
                st.plotly_chart(fig_type, use_container_width=True)
            
            # Detailed Analysis
            st.markdown('<h2 class="section-header">🔍 Detailed Evaluation Analysis by Ticket Type</h2>', unsafe_allow_html=True)
            
            # Get detailed breakdown
            breakdown_df = get_evaluation_breakdown_by_type(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
            
            if not breakdown_df.empty:
                # Calculate metrics for each ticket type
                ticket_types = breakdown_df['ticket_type'].unique()
                
                for ticket_type in ticket_types:
                    if ticket_type == 'unknown':
                        continue
                        
                    type_data = breakdown_df[breakdown_df['ticket_type'] == ticket_type]
                    total_count = type_data['count'].sum()
                    
                    if total_count > 0:
                        st.markdown(f"### {ticket_type.title()} Tickets ({total_count} total)")
                        
                        # Calculate percentages
                        copy_paste = type_data[type_data['quality'] == 'copy_paste']['count'].iloc[0] if 'copy_paste' in type_data['quality'].values else 0
                        low_quality = type_data[type_data['quality'] == 'low_quality']['count'].iloc[0] if 'low_quality' in type_data['quality'].values else 0
                        high_quality = type_data[type_data['quality'] == 'high_quality']['count'].iloc[0] if 'high_quality' in type_data['quality'].values else 0
                        skipped = type_data[type_data['quality'] == 'skipped']['count'].iloc[0] if 'skipped' in type_data['quality'].values else 0
                        
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("Copy-Paste", f"{copy_paste} ({(copy_paste/total_count*100):.1f}%)")
                        with col2:
                            st.metric("Low Quality", f"{low_quality} ({(low_quality/total_count*100):.1f}%)")
                        with col3:
                            st.metric("High Quality", f"{high_quality} ({(high_quality/total_count*100):.1f}%)")
                        with col4:
                            st.metric("Skipped", f"{skipped} ({(skipped/total_count*100):.1f}%)")
                
                # Overall Summary
                st.markdown("### Overall Summary")
                
                overall_total = breakdown_df['count'].sum()
                overall_copy_paste = breakdown_df[breakdown_df['quality'] == 'copy_paste']['count'].sum()
                overall_low_quality = breakdown_df[breakdown_df['quality'] == 'low_quality']['count'].sum()
                overall_high_quality = breakdown_df[breakdown_df['quality'] == 'high_quality']['count'].sum()
                overall_skipped = breakdown_df[breakdown_df['quality'] == 'skipped']['count'].sum()
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Tickets", f"{overall_total:,}")
                with col2:
                    st.metric("Copy-Paste %", f"{(overall_copy_paste/overall_total*100):.1f}%" if overall_total > 0 else "0%")
                with col3:
                    st.metric("Low Quality %", f"{(overall_low_quality/overall_total*100):.1f}%" if overall_total > 0 else "0%")
                with col4:
                    st.metric("Skipped %", f"{(overall_skipped/overall_total*100):.1f}%" if overall_total > 0 else "0%")
                
                # Quality breakdown chart
                fig_breakdown = create_evaluation_quality_chart(breakdown_df)
                st.plotly_chart(fig_breakdown, use_container_width=True)
                
                # Data table
                st.markdown("### Raw Data")
                st.dataframe(breakdown_df, use_container_width=True)
            
        else:
            st.warning("⚠️ No data found for the selected date range.")
            
            # Show info about available data
            min_date, max_date = get_available_dates()
            if min_date and max_date:
                st.info(f"📅 Available data range: {min_date} to {max_date}")
    
    # Tab 2: Daily Breakdown Analysis
    with tab2:
        st.markdown('<h2 class="section-header">📊 Daily Breakdown Analysis</h2>', unsafe_allow_html=True)
        st.markdown("This tab provides the same comprehensive analysis as the `create_daily_breakdown_spreadsheet.py` script.")
        
        # Create the detailed breakdown
        result = create_detailed_breakdown_spreadsheet(
            start_date.strftime('%Y-%m-%d'), 
            end_date.strftime('%Y-%m-%d')
        )
        
        if result is None:
            st.warning("⚠️ No data found for the selected date range.")
            return
        
        breakdown_df, summary_df = result
        
        # Display summary metrics
        st.markdown('<h3 class="section-header">📈 Summary Metrics</h3>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        total_records = len(breakdown_df[breakdown_df['Ticket_Type'] != 'TOTAL'])
        with col1:
            st.metric("Total Breakdown Records", f"{total_records:,}")
        
        total_summary = len(summary_df)
        with col2:
            st.metric("Total Summary Records", f"{total_summary:,}")
        
        unique_dates = len(breakdown_df[breakdown_df['Ticket_Type'] == 'TOTAL'])
        with col3:
            st.metric("Unique Dates", f"{unique_dates:,}")
        
        # Create and display the daily breakdown chart
        st.markdown('<h3 class="section-header">📊 Daily Breakdown Visualization</h3>', unsafe_allow_html=True)
        
        fig_breakdown = create_daily_breakdown_chart(breakdown_df)
        st.plotly_chart(fig_breakdown, use_container_width=True)
        
        # Display the breakdown data
        st.markdown('<h3 class="section-header">📋 Daily Breakdown Data</h3>', unsafe_allow_html=True)
        st.dataframe(breakdown_df, use_container_width=True)
        
        # Display the summary data
        st.markdown('<h3 class="section-header">📋 Summary Data</h3>', unsafe_allow_html=True)
        st.dataframe(summary_df, use_container_width=True)
        
        # Download functionality
        st.markdown('<h3 class="section-header">💾 Download Data</h3>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Download breakdown as CSV
            csv_breakdown = breakdown_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Breakdown CSV",
                data=csv_breakdown,
                file_name=f"daily_breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        with col2:
            # Download summary as CSV
            csv_summary = summary_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Summary CSV",
                data=csv_summary,
                file_name=f"summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        # Download Excel file with all sheets
        st.markdown('<h4>📊 Download Complete Excel Spreadsheet</h4>', unsafe_allow_html=True)
        st.markdown("Download the complete Excel file with multiple sheets (Daily Breakdown, Summary, Pivot Tables)")
        
        excel_file = download_excel_file(breakdown_df, summary_df)
        st.download_button(
            label="📥 Download Excel Spreadsheet",
            data=excel_file.getvalue(),
            file_name=f"daily_breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Show sample of the data
        st.markdown('<h3 class="section-header">🔍 Data Preview</h3>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Sample of Daily Breakdown Data:**")
            st.dataframe(breakdown_df.head(10), use_container_width=True)
        
        with col2:
            st.markdown("**Sample of Summary Data:**")
            st.dataframe(summary_df.head(10), use_container_width=True)
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; font-size: 0.8rem;">
        📊 Ticket Evaluation Dashboard | Powered by LangSmith & Streamlit
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()